from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook


TOKEN_PATTERN = re.compile(r"([^\.\[\]]+)|\[(\d+)\]")
SUPPORTED_SPREADSHEETS = {".csv", ".xlsx"}


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value) -> bool:
    text = normalize_text(value).lower()
    if text in {"true", "1", "si", "s", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"Valor booleano no valido: {value!r}")


def parse_number(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
    else:
        text = normalize_text(value)
        if not text:
            raise ValueError("No se puede convertir una cadena vacia a numero.")
        number = float(text)
    return int(number) if number.is_integer() else number


def coerce_value(raw_value, declared_type: str):
    kind = normalize_text(declared_type).lower() or "string"

    if kind == "string":
        return "" if raw_value is None else str(raw_value)
    if kind == "number":
        return parse_number(raw_value)
    if kind == "boolean":
        return parse_bool(raw_value)
    if kind == "null":
        return None
    if kind == "json":
        if isinstance(raw_value, (dict, list, int, float, bool)) or raw_value is None:
            return raw_value
        text = normalize_text(raw_value)
        return json.loads(text or "null")
    if kind == "empty":
        return ""

    raise ValueError(f"Tipo no soportado: {declared_type!r}")


def parse_path(path: str):
    tokens = []
    for key, index in TOKEN_PATTERN.findall(path):
        if key:
            tokens.append(key)
        else:
            tokens.append(int(index))
    if not tokens:
        raise ValueError(f"Ruta no valida: {path!r}")
    return tokens


def ensure_list_size(target: list, index: int):
    while len(target) <= index:
        target.append(None)


def set_nested_value(root: dict, path: str, value):
    tokens = parse_path(path)
    cursor = root

    for position, token in enumerate(tokens):
        is_last = position == len(tokens) - 1
        next_token = None if is_last else tokens[position + 1]

        if isinstance(token, str):
            if is_last:
                cursor[token] = value
                return

            if token not in cursor or cursor[token] is None:
                cursor[token] = [] if isinstance(next_token, int) else {}
            cursor = cursor[token]
            continue

        ensure_list_size(cursor, token)
        if is_last:
            cursor[token] = value
            return

        if cursor[token] is None:
            cursor[token] = [] if isinstance(next_token, int) else {}
        cursor = cursor[token]


def read_csv_rows(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if not row:
                continue
            if not any(normalize_text(value) for value in row.values()):
                continue
            yield {
                "period": row.get("period"),
                "path": row.get("path"),
                "type": row.get("type"),
                "value": row.get("value"),
            }


def read_xlsx_rows(path: Path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["updates"] if "updates" in workbook.sheetnames else workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return

    headers = [normalize_text(value).lower() for value in values[0]]
    required = ["period", "path", "type", "value"]
    positions = {name: headers.index(name) for name in required}

    for row in values[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        yield {
            "period": row[positions["period"]] if positions["period"] < len(row) else None,
            "path": row[positions["path"]] if positions["path"] < len(row) else None,
            "type": row[positions["type"]] if positions["type"] < len(row) else None,
            "value": row[positions["value"]] if positions["value"] < len(row) else None,
        }


def spreadsheet_to_json(input_path: Path, output_path: Path):
    if input_path.suffix.lower() not in SUPPORTED_SPREADSHEETS:
        raise ValueError(f"Formato no soportado: {input_path.suffix}")

    rows = read_xlsx_rows(input_path) if input_path.suffix.lower() == ".xlsx" else read_csv_rows(input_path)
    result = {}

    for index, row in enumerate(rows, start=2):
        period = normalize_text(row["period"])
        path = normalize_text(row["path"])
        declared_type = normalize_text(row["type"])

        if not period or not path:
            raise ValueError(f"Fila {index}: period y path son obligatorios.")

        if period not in result:
            result[period] = {}

        value = coerce_value(row["value"], declared_type)
        set_nested_value(result[period], path, value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def infer_type(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    return "string"


def flatten_value(period: str, path: str, value, rows: list[dict]):
    if isinstance(value, dict):
        if not value:
            rows.append(
                {"period": period, "path": path, "type": "json", "value": "{}"}
            )
            return
        for key, child in value.items():
            child_path = key if not path else f"{path}.{key}"
            flatten_value(period, child_path, child, rows)
        return

    if isinstance(value, list):
        if not value:
            rows.append(
                {"period": period, "path": path, "type": "json", "value": "[]"}
            )
            return
        for index, child in enumerate(value):
            child_path = f"{path}[{index}]"
            flatten_value(period, child_path, child, rows)
        return

    rows.append(
        {
            "period": period,
            "path": path,
            "type": infer_type(value),
            "value": "" if value is None else value,
        }
    )


def json_to_csv(input_path: Path, output_path: Path):
    data = json.loads(input_path.read_text(encoding="utf-8"))
    rows: list[dict] = []

    for period in sorted(data.keys()):
        flatten_value(period, "", data[period], rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["period", "path", "type", "value"])
        writer.writeheader()
        writer.writerows(rows)


def build_parser():
    parser = argparse.ArgumentParser(
        description="Convierte un archivo de Excel/CSV a JSON para el boletin, o genera la plantilla CSV."
    )
    parser.add_argument("source", help="Ruta del archivo origen.")
    parser.add_argument("target", help="Ruta del archivo destino.")
    parser.add_argument(
        "--from-json",
        action="store_true",
        help="Genera un CSV editable a partir de un JSON existente.",
    )
    return parser


def main():
    args = build_parser().parse_args()
    source = Path(args.source)
    target = Path(args.target)

    if args.from_json:
        json_to_csv(source, target)
        return

    spreadsheet_to_json(source, target)


if __name__ == "__main__":
    main()
