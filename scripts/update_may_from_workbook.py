from __future__ import annotations

import copy
import json
import math
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
JSON_PATH = REPO / "public" / "data" / "boletin-data.json"
MONTHS = {
    1: ("ENERO", "Enero", "Ene"),
    2: ("FEBRERO", "Febrero", "Feb"),
    3: ("MARZO", "Marzo", "Mar"),
    4: ("ABRIL", "Abril", "Abr"),
    5: ("MAYO", "Mayo", "May"),
}
PORTFOLIO_COLORS = ["#1967e8", "#28105f", "#ea2c20", "#f1d554", "#ef7b7b"]


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def money_m(value, decimals=1, thousands=True):
    value = value if is_number(value) else 0
    amount = value / 1_000_000
    text = f"{amount:,.{decimals}f}" if thousands else f"{amount:.{decimals}f}"
    text = text.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"${text} M"


def integer(value):
    value = value if is_number(value) else 0
    return f"{int(round(value)):,}".replace(",", ".")


def percent(value, decimals=1, signed=False, spaced=False):
    value = value if is_number(value) else 0
    sign = "+" if signed and value > 0 else ""
    gap = " " if spaced else ""
    return f"{sign}{value * 100:.{decimals}f}{gap}%".replace(".", ",")


def compact(value):
    value = value if is_number(value) else 0
    if abs(value) >= 1_000_000:
        return f"{value / 1_000_000:.2f} M".replace(".", ",")
    if abs(value) >= 1_000:
        return f"{value / 1_000:.2f} mil".replace(".", ",")
    return f"{value:.0f}"


def money_compact(value):
    value = value if is_number(value) else 0
    if abs(value) >= 1_000_000_000:
        return f"${value / 1_000_000_000:.2f} mil M".replace(".", ",")
    if abs(value) >= 1_000_000:
        return f"${value / 1_000_000:.2f} M".replace(".", ",")
    if abs(value) >= 1_000:
        return f"${value / 1_000:.2f} mil".replace(".", ",")
    return f"${value:.0f}"


def tone_target(value):
    if not is_number(value):
        return "red"
    if value > 0.995:
        return "green"
    if value > 0.895:
        return "yellow"
    return "red"


def tone_delta(value):
    return "green" if is_number(value) and value >= 0 else "red"


def by_prefix(values, prefix):
    for key, value in values.items():
        if key.startswith(prefix):
            return value
    return None


def sheet(wb, expected):
    for name in wb.sheetnames:
        if name.lower() == expected.lower():
            return wb[name]
    raise KeyError(expected)


def all_rows(ws):
    return [list(row) for row in ws.iter_rows(values_only=True)]


def month_col(header, month):
    for idx, value in enumerate(header):
        if isinstance(value, datetime) and value.year == 2026 and value.month == month:
            return idx
    raise KeyError(month)


def block(rows, title, month):
    start = next(i for i, row in enumerate(rows) if row and row[0] == title)
    col = month_col(rows[start], month)
    result = {}
    for row in rows[start + 1 :]:
        label = row[0] if row else None
        if label in (None, ""):
            continue
        if isinstance(label, str) and any(isinstance(value, datetime) for value in row):
            break
        if isinstance(label, str) and label.strip().lower() in {
            "margenes",
            "cumplimiento meta",
            "cumplimiento meta",
            "var. mes anterior",
            "var. vr mes anterior",
            "var. abs. vrs meta",
        }:
            break
        result[str(label).strip()] = row[col] if col < len(row) else None
    return result


def trans_block(rows, title, month):
    return block(rows, title, month)


def update_ingresos(period, wb, month):
    prev = month - 1
    short_prev = MONTHS[prev][2].lower()
    ingreso_rows = all_rows(sheet(wb, "Ingresos 1"))
    result_rows = all_rows(sheet(wb, "Resultados 1"))
    real = block(ingreso_rows, "Ingresos real", month)
    meta = block(ingreso_rows, "Meta", month)
    cumplimiento = block(ingreso_rows, "Cumplimiento Meta", month)
    variacion = block(ingreso_rows, "Var. Mes anterior", month)
    prev_real = block(ingreso_rows, "Ingresos real", prev)
    prev_meta = block(ingreso_rows, "Meta", prev)
    result_real = block(result_rows, "Resultados ingresos", month)
    prev_result_real = block(result_rows, "Resultados ingresos", prev)
    result_cump = block(result_rows, "Cumplimiento Meta", month)

    def result_variation(prefix):
        current = by_prefix(result_real, prefix)
        previous = by_prefix(prev_result_real, prefix)
        return (current - previous) / previous if previous else None

    card_source = {
        "ePayco pagos Agregador": {
            "real": by_prefix(result_real, "Ingresos Agreador") or real.get("ePayco pagos Agregador"),
            "cumplimiento": cumplimiento.get("ePayco pagos Agregador"),
            "variacion": result_variation("Ingresos Agreador"),
        },
        "ePayco pagos Gateway": {
            "real": by_prefix(result_real, "Ingresos Gateway") or real.get("ePayco pagos Gateway"),
            "cumplimiento": cumplimiento.get("ePayco pagos Gateway"),
            "variacion": result_variation("Ingresos Gateway"),
        },
        "__otras__": {
            "real": by_prefix(result_real, "Ingresos otros"),
            "cumplimiento": by_prefix(result_cump, "Ingresos otros"),
            "variacion": result_variation("Ingresos otros"),
        },
        "Ingresos total": {
            "real": by_prefix(result_real, "Total ingresos") or real.get("Ingresos total"),
            "cumplimiento": cumplimiento.get("Ingresos total"),
            "variacion": result_variation("Total ingresos"),
        },
    }

    cards = []
    for title, key, subtitle in [
        ("Ingresos Agregador", "ePayco pagos Agregador", None),
        ("Ingresos Gateway", "ePayco pagos Gateway", "Afiliaciones + transacciones"),
        ("Otras líneas", "__otras__", None),
        ("Ingreso total", "Ingresos total", None),
    ]:
        values = card_source.get(key, {})
        real_value = values.get("real", real.get(key))
        cump_value = values.get("cumplimiento", cumplimiento.get(key))
        var_value = values.get("variacion", variacion.get(key))
        card = {
            "title": title,
            "value": money_m(real_value),
            "delta": f"{percent(var_value, signed=True, spaced=True)} vs {short_prev}",
            "meta": f"{percent(cump_value)} de meta",
            "deltaTone": tone_delta(var_value),
            "metaTone": tone_target(cump_value),
        }
        if subtitle:
            card["subtitle"] = subtitle
        cards.append(card)
    period["ingresos"]["cards"] = cards
    period["ingresos"]["summary"] = {
        "total": money_m(real.get("Ingresos total"), thousands=False),
        "delta": percent(variacion.get("Ingresos total"), signed=True),
        "cumplimiento": f"Cumpl. {percent(cumplimiento.get('Ingresos total'))}",
        "ene": money_m(prev_real.get("Ingresos total")),
        "ppto": money_m(meta.get("Ingresos total")),
    }

    margin_real = block(result_rows, "Margen real", month)
    margin_meta = block(result_rows, "Margen Meta", month)
    margin_var = block(result_rows, "Var. Vr mes anterior", month)
    prev_margin_real = block(result_rows, "Margen real", prev)
    margins = []
    for title, key in [
        ("Margen bruto", "Margen Bruto"),
        ("Margen EBITDA", "Margen Operacional"),
        ("Margen neto", "Utilidad neta"),
    ]:
        real_value = margin_real.get(key)
        meta_value = margin_meta.get(key)
        cump_value = real_value / meta_value if meta_value else None
        var_value = margin_var.get(key)
        margins.append(
            {
                "title": title,
                "real": money_m(real_value),
                "ppto": money_m(meta_value),
                "pill": percent(cump_value),
                "pillTone": tone_target(cump_value),
                "vsText": percent(var_value, signed=True),
                "vsTone": tone_delta(var_value),
                "note": f"Variación absoluta vs meta: {money_m((real_value or 0) - (meta_value or 0))}",
                "series": [
                    {"name": MONTHS[prev][2].upper(), "value": int((prev_margin_real.get(key) or 0) / 100000)},
                    {"name": f"{MONTHS[month][2].upper()} REAL", "value": int((real_value or 0) / 100000)},
                    {"name": "PPTO", "value": int((meta_value or 0) / 100000)},
                ],
            }
        )
    period["ingresos"]["margins"] = margins

    names = [
        "Ingresos total",
        "ePayco pagos Agregador",
        "ePayco pagos Gateway",
        "ePayco Recaudo",
        "Suscripciones",
        "ePayco Control",
        "ePayco Paypal",
        "ePayco shops",
        "ePayco PayOuts",
    ]
    table = []
    for name in names:
        pr, pm = prev_real.get(name), prev_meta.get(name)
        cr, cm = real.get(name), meta.get(name)
        table.append(
            [
                name,
                money_m(pr),
                money_m(pm),
                percent(pr / pm if pm else None),
                money_m(cr),
                money_m(cm),
                percent(cr / cm if cm else None),
                percent((cr - pr) / pr if pr else None, signed=True),
            ]
        )
    period["ingresos"]["table"] = table


def update_transaccionales(period, wb, month):
    rows = all_rows(sheet(wb, "Transaccionales"))
    prev = month - 1
    names = ["Portafolio Movistar", "Portafolio Wom", "Portafolio Davivienda", "Portafolio General", "Portafolio especial"]
    values = trans_block(rows, "Real Trx Gateway x negociación", month)
    shares = trans_block(rows, "Real % Part. Trx x negociación", month)
    period["transaccionales"]["pie"] = [
        {"name": name, "value": round((shares.get(name) or 0) * 100, 1), "color": PORTFOLIO_COLORS[index]}
        for index, name in enumerate(names)
    ]
    period["transaccionales"]["values"] = [[name, integer(values.get(name))] for name in names]
    period["transaccionales"]["total"] = integer(values.get("Total"))

    def make_card(title, real, meta, cump, prev_real, prev_meta, money=False):
        return {
            "title": title,
            "real": money_compact(real) if money else compact(real),
            "meta": money_compact(meta) if money else compact(meta),
            "cumplimiento": percent(cump, decimals=0),
            "realPrev": money_compact(prev_real) if money else compact(prev_real),
            "metaPrev": money_compact(prev_meta) if money else compact(prev_meta),
            "tone": tone_target(cump),
        }

    agg_real = trans_block(rows, "KPIS -Agregador Real", month)
    agg_meta = trans_block(rows, "KPIS -Agregador Supuestos", month)
    agg_cump = trans_block(rows, "KPIS -Agregador Cumplimiento", month)
    agg_prev_real = trans_block(rows, "KPIS -Agregador Real", prev)
    agg_prev_meta = trans_block(rows, "KPIS -Agregador Supuestos", prev)
    period["vinculacion"]["kpiAgregador"] = [
        make_card("#Transacciones", agg_real.get("# Transaciones"), agg_meta.get("# Transaciones"), agg_cump.get("# Transaciones"), agg_prev_real.get("# Transaciones"), agg_prev_meta.get("# Transaciones")),
        make_card("$ Dinero operado", agg_real.get("$Dinero operado"), agg_meta.get("$Dinero operado"), agg_cump.get("$Dinero operado"), agg_prev_real.get("$Dinero operado"), agg_prev_meta.get("$Dinero operado"), True),
        make_card("Ticket promedio", agg_real.get("Ticket promedio"), agg_meta.get("Ticket promedio"), agg_cump.get("Ticket promedio"), agg_prev_real.get("Ticket promedio"), agg_prev_meta.get("Ticket promedio"), True),
        make_card("Clientes transaccionales", agg_real.get("Clientes transaccionales"), agg_meta.get("Clientes transaccionales"), agg_cump.get("Clientes transaccionales"), agg_prev_real.get("Clientes transaccionales"), agg_prev_meta.get("Clientes transaccionales")),
    ]

    gw_real = trans_block(rows, "Real. KPIS - Gateway", month)
    gw_meta = trans_block(rows, "Proy. KPIS - Gateway", month)
    gw_cump = trans_block(rows, "KPIS -Gateway Cumplimiento", month)
    gw_prev_real = trans_block(rows, "Real. KPIS - Gateway", prev)
    gw_prev_meta = trans_block(rows, "Proy. KPIS - Gateway", prev)
    period["vinculacion"]["kpiGateway"] = [
        make_card("#Transacciones", gw_real.get("# Transaciones"), gw_meta.get("# Transaciones"), gw_cump.get("# Transaciones"), gw_prev_real.get("# Transaciones"), gw_prev_meta.get("# Transaciones")),
        make_card("Clientes", gw_real.get("Clientes"), gw_meta.get("Clientes"), gw_cump.get("Clientes"), gw_prev_real.get("Clientes"), gw_prev_meta.get("Clientes")),
        make_card("TXR promedio x cliente", gw_real.get("TRX promedio  x cliente"), gw_meta.get("TRX promedio  x cliente"), gw_cump.get("TRX promedio  x cliente"), gw_prev_real.get("TRX promedio  x cliente"), gw_prev_meta.get("TRX promedio  x cliente")),
    ]


def update_churn(period, wb):
    ws = sheet(wb, "Churn")
    by_month_agregador = {}
    by_month_gateway = {}
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        if len(row) > 12 and row[8] in {"Ene", "Feb", "Mar", "Abr", "May"}:
            by_month_agregador[row[8]] = row
        if len(row) > 19 and row[15] in {"Ene", "Feb", "Mar", "Abr", "May"}:
            by_month_gateway[row[15]] = row

    def apply_dataset(dataset, source, offsets):
        activos, churn, nuevos = [], [], []
        month_col, activos_col, nuevos_col, churn_col, rate_col = offsets
        for name in ["Ene", "Feb", "Mar", "Abr", "May"]:
            row = source.get(name)
            if not row or not is_number(row[activos_col]):
                continue
            activos.append({"name": row[month_col], "value": int(row[activos_col])})
            nuevos.append({"name": row[month_col], "value": int(row[nuevos_col])})
            churn.append(
                {
                    "name": row[month_col],
                    "value": int(row[churn_col]),
                    "rate": percent(row[rate_col], decimals=2),
                }
            )
        if not activos:
            return
        dataset["activos"] = activos
        dataset["nuevos"] = nuevos
        dataset["churn"] = churn
        dataset["promedioActivos"] = integer(sum(item["value"] for item in activos) / len(activos))
        dataset["promedioNuevos"] = integer(sum(item["value"] for item in nuevos) / len(nuevos))
        dataset["promedioChurn"] = integer(sum(item["value"] for item in churn) / len(churn))
        dataset["tasaMes"] = churn[-1]["rate"]

    apply_dataset(period["transaccionales"]["churnAgregador"], by_month_agregador, (8, 9, 10, 11, 12))
    apply_dataset(period["transaccionales"]["churnGateway"], by_month_gateway, (15, 16, 17, 18, 19))


def update_nps(period, wb, month):
    rows = all_rows(sheet(wb, "NPS"))
    headers = [str(value) if value is not None else "" for value in rows[3]]
    col = headers.index(f"2026-{month}")
    prev_col = headers.index(f"2026-{month - 1}")
    values = {row[0]: row for row in rows if row and row[0]}

    def metric(cump_label, value_label, color, suffix="", reverse=False):
        value = values[value_label][col]
        prev = values[value_label][prev_col]
        trend = "↓" if is_number(prev) and is_number(value) and value < prev else "↑"
        tone = "down" if trend == "↓" else "up"
        if reverse:
            tone = "up" if trend == "↓" else "down"
        return {
            "label": value_label.replace("Tiempo Espera Promedio (", "").replace("Tiempo Duración Promedio (", "").replace(")", ""),
            "value": (f"{value:.1f}{suffix}" if is_number(value) else str(value)).replace(".", ","),
            "trend": trend,
            "trendTone": tone,
            "progress": int(round((values[cump_label][col] or 0) * 100)),
            "color": color,
        }

    period["nps"]["months"] = [
        period["nps"]["months"][-1],
        {
            "month": MONTHS[month][1],
            "code": f"2026-{month}",
            "insight": "NPS y SSA mejoran frente al mes anterior. TEP se mantiene por encima de la meta; TDP continúa como el principal punto de atención.",
            "metrics": [
                metric("% Cumplimiento NPS", "NPS", "#ef5a5a"),
                metric("% Cumplimiento SSA", "SSA", "#15516f"),
                metric("% Cumplimiento TEP", "Tiempo Espera Promedio (TEP)", "#20a751", "min", True),
                metric("% Cumplimiento TDP", "Tiempo Duración Promedio (TDP)", "#f1a10a", "h", True),
            ],
        },
    ]


def update_vinculacion(period, wb, month):
    rows = all_rows(sheet(wb, "vinculaciones"))
    cols = []
    for current in range(1, month + 1):
        for idx, value in enumerate(rows[2]):
            if isinstance(value, datetime) and value.year == 2026 and value.month == current:
                cols.append(idx)
                break
    period["vinculacion"]["months"] = [
        {
            "label": f"{MONTHS[idx + 1][2]}-26",
            "fact": f"Fact. {money_m(rows[7][col])}",
            "activacion": percent(rows[6][col], decimals=0),
            "registros": integer(rows[3][col]),
            "vinc": integer(rows[4][col]),
            "ccios": integer(rows[5][col]),
        }
        for idx, col in enumerate(cols)
    ]


def main(workbook_path: str):
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    for month in [4, 5]:
        key = f"2026-{month:02d}"
        period = copy.deepcopy(data.get(key) or data[f"2026-{month - 1:02d}"])
        period["corteMes"] = MONTHS[month][0]
        period["corteAnio"] = "2026"
        update_ingresos(period, wb, month)
        update_transaccionales(period, wb, month)
        update_churn(period, wb)
        update_nps(period, wb, month)
        update_vinculacion(period, wb, month)
        data[key] = period
    JSON_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("updated", ", ".join(sorted(data.keys())))


if __name__ == "__main__":
    import sys

    main(sys.argv[1])
